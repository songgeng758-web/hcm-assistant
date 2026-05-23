"""
生成员工档案标准模板 Excel
运行方式（在 backend 目录下）：python -m app.utils.gen_template
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings


COLUMNS = [
    ("工号", "必填，员工唯一编号", "E0001"),
    ("姓名", "必填", "张三"),
    ("身份证号", "必填，18位", "110101199003078812"),
    ("性别", "男 / 女，会与身份证比对", "男"),
    ("手机号", "11位手机号", "13800138000"),
    ("邮箱", "选填", "zhangsan@example.com"),
    ("部门", "所属部门", "技术部"),
    ("岗位", "职位名称", "软件工程师"),
    ("入职日期", "YYYY-MM-DD", "2024-01-15"),
]


def generate(output_path: Path):
    df = pd.DataFrame([{col: example for col, _, example in COLUMNS}])
    df.to_excel(output_path, index=False, sheet_name="员工档案")

    wb = load_workbook(output_path)
    ws = wb["员工档案"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    align_center = Alignment(horizontal="center", vertical="center")

    for col_idx, (col_name, desc, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) * 2 + 4, 14)

    wb.save(output_path)

    desc_df = pd.DataFrame(
        [{"字段名": c, "说明": d, "示例": e} for c, d, e in COLUMNS]
    )
    with pd.ExcelWriter(output_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        desc_df.to_excel(writer, sheet_name="字段说明", index=False)

    print(f"模板已生成: {output_path}")


if __name__ == "__main__":
    output = settings.TEMPLATE_DIR / "employee_template.xlsx"
    generate(output)